from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from backend.app import db
from backend.models.producto import Producto
from backend.models.transaccion import Transaccion
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

excel_bp = Blueprint('excel', __name__)

def aplicar_estilos_encabezado(ws, columnas):
    """Aplicar estilos a los encabezados"""
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, columna in enumerate(columnas, 1):
        celda = ws.cell(row=1, column=col_num)
        celda.fill = fill
        celda.font = font
        celda.alignment = alignment
        celda.border = border

def aplicar_estilos_datos(ws, fila_inicio, fila_fin, num_columnas):
    """Aplicar estilos a las filas de datos"""
    alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for fila in range(fila_inicio, fila_fin + 1):
        for col in range(1, num_columnas + 1):
            celda = ws.cell(row=fila, column=col)
            celda.border = border
            celda.alignment = alignment

@excel_bp.route('/exportar-inventario', methods=['GET'])
@jwt_required()
def exportar_inventario():
    """Exportar inventario a Excel"""
    productos = Producto.query.filter_by(estado='activo').all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Encabezados
    columnas = ['Código', 'Nombre', 'Descripción', 'Categoría', 'Proveedor', 
                'Cantidad', 'Cantidad Mínima', 'Precio Unitario', 'Valor Total', 
                'Bajo Stock', 'Fecha Ingreso']
    
    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna)
    
    aplicar_estilos_encabezado(ws, columnas)
    
    # Datos
    for fila_num, producto in enumerate(productos, 2):
        ws.cell(row=fila_num, column=1, value=producto.codigo)
        ws.cell(row=fila_num, column=2, value=producto.nombre)
        ws.cell(row=fila_num, column=3, value=producto.descripcion)
        ws.cell(row=fila_num, column=4, value=producto.categoria)
        ws.cell(row=fila_num, column=5, value=producto.proveedor)
        ws.cell(row=fila_num, column=6, value=producto.cantidad)
        ws.cell(row=fila_num, column=7, value=producto.cantidad_minima)
        ws.cell(row=fila_num, column=8, value=producto.precio_unitario)
        ws.cell(row=fila_num, column=9, value=producto.cantidad * producto.precio_unitario)
        ws.cell(row=fila_num, column=10, value="Sí" if producto.cantidad < producto.cantidad_minima else "No")
        ws.cell(row=fila_num, column=11, value=producto.fecha_ingreso.strftime('%Y-%m-%d'))
    
    aplicar_estilos_datos(ws, 2, len(productos) + 1, len(columnas))
    
    # Ajustar ancho de columnas
    for col in range(1, len(columnas) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@excel_bp.route('/exportar-transacciones', methods=['GET'])
@jwt_required()
def exportar_transacciones():
    """Exportar transacciones a Excel"""
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    query = Transaccion.query
    
    if fecha_inicio:
        fecha_inicio_obj = datetime.fromisoformat(fecha_inicio)
        query = query.filter(Transaccion.fecha_transaccion >= fecha_inicio_obj)
    
    if fecha_fin:
        fecha_fin_obj = datetime.fromisoformat(fecha_fin)
        query = query.filter(Transaccion.fecha_transaccion <= fecha_fin_obj)
    
    transacciones = query.order_by(Transaccion.fecha_transaccion.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    
    # Encabezados
    columnas = ['Fecha', 'Producto', 'Código', 'Tipo', 'Cantidad', 'Usuario', 
                'Motivo', 'Referencia', 'Observaciones']
    
    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna)
    
    aplicar_estilos_encabezado(ws, columnas)
    
    # Datos
    for fila_num, trans in enumerate(transacciones, 2):
        ws.cell(row=fila_num, column=1, value=trans.fecha_transaccion.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=fila_num, column=2, value=trans.producto.nombre if trans.producto else '')
        ws.cell(row=fila_num, column=3, value=trans.producto.codigo if trans.producto else '')
        ws.cell(row=fila_num, column=4, value=trans.tipo.upper())
        ws.cell(row=fila_num, column=5, value=trans.cantidad)
        ws.cell(row=fila_num, column=6, value=trans.usuario_registrador.usuario if trans.usuario_registrador else '')
        ws.cell(row=fila_num, column=7, value=trans.motivo)
        ws.cell(row=fila_num, column=8, value=trans.referencia)
        ws.cell(row=fila_num, column=9, value=trans.observaciones)
    
    aplicar_estilos_datos(ws, 2, len(transacciones) + 1, len(columnas))
    
    # Ajustar ancho de columnas
    for col in range(1, len(columnas) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'transacciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@excel_bp.route('/exportar-bajo-stock', methods=['GET'])
@jwt_required()
def exportar_bajo_stock():
    """Exportar productos con bajo stock a Excel"""
    productos = Producto.query.filter(
        Producto.cantidad < Producto.cantidad_minima,
        Producto.estado == 'activo'
    ).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bajo Stock"
    
    # Encabezados
    columnas = ['Código', 'Nombre', 'Cantidad Actual', 'Cantidad Mínima', 
                'Faltante', 'Proveedor', 'Precio Unitario']
    
    for col_num, columna in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=columna)
    
    aplicar_estilos_encabezado(ws, columnas)
    
    # Datos
    for fila_num, producto in enumerate(productos, 2):
        ws.cell(row=fila_num, column=1, value=producto.codigo)
        ws.cell(row=fila_num, column=2, value=producto.nombre)
        ws.cell(row=fila_num, column=3, value=producto.cantidad)
        ws.cell(row=fila_num, column=4, value=producto.cantidad_minima)
        ws.cell(row=fila_num, column=5, value=producto.cantidad_minima - producto.cantidad)
        ws.cell(row=fila_num, column=6, value=producto.proveedor)
        ws.cell(row=fila_num, column=7, value=producto.precio_unitario)
    
    aplicar_estilos_datos(ws, 2, len(productos) + 1, len(columnas))
    
    # Ajustar ancho de columnas
    for col in range(1, len(columnas) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'bajo_stock_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )